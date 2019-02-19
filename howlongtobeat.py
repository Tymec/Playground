from hltbScraper import HLTB
import openpyxl
import argparse

parser = argparse.ArgumentParser(description='Files from a path to paths in json file')
parser.add_argument('-i', '--xlsx', help="Path to the xlsx file", dest="xlsx", metavar="xlsx", required=True)
parser.add_argument('-o', '--out', help="Path to the output file", dest="out", metavar="out", required=True)
parser.add_argument('-s', '--sheet', help="Name of the sheet to use", dest="sheet", metavar="sheet", required=True)
args = vars(parser.parse_args())

wb = openpyxl.load_workbook(args["xlsx"])
sheet = wb.get_sheet_by_name(args["sheet"])

flush = ""
for i in range(0, 100):
    flush += " "

i = 1
while True:
    i += 1
    _ignore = sheet.cell(row=i, column=1).value
    cell = sheet.cell(row=i, column=2).value
    if cell is None:
        break
    elif _ignore is "*":
        continue
    try:
        game = HLTB(cell)
        time_to_beat = next(iter(game.values()))['Main Story']
        
        if "--" in time_to_beat:
            time_to_beat = "Not Found"
    except IndexError:
        time_to_beat = "Not Found"
    except Exception as e:
        if "Not Found" in str(e):
            time_to_beat = "Not Found"
    sheet['D' + str(i)] = time_to_beat
    
    print(flush, end="\r")
    print("#{}: {} - {}".format(i, cell, time_to_beat), end="\r")
print("")
wb.save(args["xlsx"])
